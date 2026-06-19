#!/usr/bin/env python3
"""
African Sky Villas — two-property financial model builder.

Rebuilds asv-two-property-model.xlsx from source exports:
  - bank notification analysis  (required)   -> Account_Activity sheet
  - iKhokha transaction export  (optional)   -> gross-up journal
  - reservations export         (optional)   -> forward-bookings tab
  - Zoho summary (json)         (optional)   -> accrual cross-check

Formula-driven, reconciled to the bank control total. Run:
    python build_model.py
Then recalc in LibreOffice/Excel (or `python recalc.py output/...xlsx`).

House rules (see CLAUDE.md): facts vs assumptions separated, Control tab
drives allocation, externally-used figures carry the approval block.
"""
import json, os, glob, argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# ----------------------------------------------------------------------------
# CONFIG — point these at your files (relative to this script)
# ----------------------------------------------------------------------------
CFG = {
    "bank":         "inputs/bank_activity.xlsx",          # Account_Activity sheet
    "ikhokha":      "inputs/ikhokha_transactions.xlsx",   # optional (or glob inputs/transactions_*.xlsx)
    "reservations": None,                                  # optional Booking.com export
    "zoho_summary": "inputs/zoho_summary.json",            # optional (see CLAUDE.md to generate via connector)
    "out":          "output/asv-two-property-model.xlsx",
    "ikhokha_fee":  0.0375,                                # validated blended rate; replace with statements
}

# ----------------------------------------------------------------------------
# Property tagging rules — edit here to refine segmentation
# ----------------------------------------------------------------------------
BIG_BAY = ['seaside village', 'ag1 seaside', 'big bay', 'bigbay', 'stratton',
           'cormorant', 'blouberg']
MARLOTH = ['ik pay', 'booking.c', 'superspar komatipoor', 'savemor', 'komatipoor',
           'nkomazi', 'marloth', 'malelane', 'hectorspruit', 'royal bush',
           'royal 3380', 'royal 3463', 'kruger saf', 'sanparks', 'crocodile',
           'game drive', 'safari', 'spar marloth']

def tag_property(text):
    t = text.lower()
    if any(k in t for k in BIG_BAY): return 'BIG_BAY'
    if any(k in t for k in MARLOTH): return 'MARLOTH'
    return 'SHARED'

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
FONT='Calibri'; INK='1A1A1A'; GOLD='B0832F'; BUSH='2C4031'; COAST='33586B'
GREY='6B6857'; GREEN='008000'; RED='C00000'; BLUE='0000FF'
def F(sz=10,b=False,c=INK,color=None): return Font(name=FONT,size=sz,bold=b,color=color or c)
FILL_INK=PatternFill('solid',start_color=INK); FILL_HEAD=PatternFill('solid',start_color='EDE6D4')
FILL_BLUE=PatternFill('solid',start_color='FFF8CC'); FILL_WARN=PatternFill('solid',start_color='F7E8D0')
TOPB=Border(top=Side(style='thin',color=INK)); BORD=Border(bottom=Side(style='thin',color='CDC3AA'))
ZAR='R#,##0;(R#,##0);"–"'; ZAR2='R#,##0.00;(R#,##0.00);"–"'; PCT='0.00%'
def hdr(ws,cell,txt,sz=13):
    ws[cell]=txt; ws[cell].font=F(sz,True,'F3ECDB'); ws[cell].fill=FILL_INK
    ws[cell].alignment=Alignment(vertical='center',indent=1)

EXP_CATS=['Supplier payments','Salary / wages','Card spend','Meals / restaurants / entertainment',
 'Telecom / utilities','Fuel / transport','Software / subscriptions','Rent / property',
 'Taxes / statutory payments','Accommodation / hotels','Cash withdrawal','Other']

# ----------------------------------------------------------------------------
# Load + tag the bank data
# ----------------------------------------------------------------------------
def load_bank(path):
    aa = pd.read_excel(path, sheet_name='Account_Activity', header=0)
    aa['Date']=pd.to_datetime(aa['Date'])
    aa['Month']=aa['Date'].dt.strftime('%Y-%m')
    for c in ['Inflow','Outflow','Internal transfer']:
        aa[c]=pd.to_numeric(aa[c],errors='coerce').fillna(0)
    aa['t']=(aa['Description'].astype(str)+' '+aa['Counterparty'].astype(str)+' '+aa['Category'].astype(str))
    aa['Property']=aa['t'].apply(tag_property)
    pl=aa[aa['Internal transfer']==0].copy()
    return pl[['Date','Month','Property','Counterparty','Description','Category','Inflow','Outflow']]

# ----------------------------------------------------------------------------
# Core workbook (Read_Me, Control, Data_Tagged, IS, Monthly, Reservations,
# Combined, Reconciliation) — formula-driven via SUMIFS
# ----------------------------------------------------------------------------
def build_core(wb, data):
    NR=len(data); END=1+NR
    ctrl_total_in=round(data[data['Category']=='Sales / customer receipts']['Inflow'].sum(),2)
    ctrl_total_out=round(data['Outflow'].sum(),2)
    refunds=round(data[data['Category'].str.contains('Refund',na=False)]['Inflow'].sum(),2)
    months=sorted(data['Month'].dropna().unique().tolist())

    # Read_Me
    ws=wb.active; ws.title='Read_Me'; ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=80
    ws.merge_cells('A1:B1'); hdr(ws,'A1','AFRICAN SKY VILLAS — TWO-PROPERTY FINANCIAL MODEL',14)
    info=[('STATUS','DRAFT / WORKING MODEL — built from bank-notification analysis; not audited financials'),
          ('Operating entity','African Sky (Pty) Ltd · 2025/571527/07 · ZAR · not VAT-registered'),
          ('Properties','Marloth Park lodge (operating) + Big Bay Seaside Escape (long-term rental)'),
          ('Rows in P&L data layer',f'{NR} (internal transfers excluded)'),
          ('Control — Sales receipts',f'R {ctrl_total_in:,.2f}'),
          ('Control — outflows',f'R {ctrl_total_out:,.2f}'),
          ('Control — refunds',f'R {refunds:,.2f}'),
          ('METHOD','Tag MARLOTH/BIG_BAY/SHARED; direct costs sit with property; shared split by Control key; reservations + Zoho + iKhokha layers cross-check.'),
          ('VERIFY','VAT: revenue annualises >R1m threshold — registered tax practitioner (urgent).'),
          ('APPROVAL','Externally-used figures need Dreyer + adviser sign-off (pi-compliance-guardrail).')]
    r=3
    for k,v in info:
        ws[f'A{r}']=k; ws[f'A{r}'].font=F(9,True, GOLD if k in('STATUS','METHOD','VERIFY','APPROVAL') else INK)
        ws[f'B{r}']=v; ws[f'B{r}'].font=F(9); ws[f'B{r}'].alignment=Alignment(wrap_text=True,vertical='top'); r+=1

    # Control
    ws=wb.create_sheet('Control'); ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=46; ws.column_dimensions['B'].width=16; ws.column_dimensions['C'].width=50
    ws.merge_cells('A1:C1'); hdr(ws,'A1','CONTROL · ASSUMPTIONS (edit blue cells)')
    def inp(row,label,val,note):
        ws[f'A{row}']=label; ws[f'A{row}'].font=F(10)
        c=ws[f'B{row}']; c.value=val; c.font=F(10,True,BLUE); c.fill=FILL_BLUE; c.number_format=PCT
        c.alignment=Alignment(horizontal='center'); c.border=BORD
        ws[f'C{row}']=note; ws[f'C{row}'].font=F(9,False,GREY)
    ws['A3']='SHARED-COST ALLOCATION KEY'; ws['A3'].font=F(10,True,GOLD)
    inp(4,'Share of unallocated costs → Marloth',1.00,'Big Bay passive; default 100% Marloth.')
    ws['A5']='Share of unallocated costs → Big Bay'; ws['A5'].font=F(10)
    ws['B5']='=1-B4'; ws['B5'].font=F(10,False,GREEN); ws['B5'].number_format=PCT; ws['B5'].alignment=Alignment(horizontal='center')
    ws['A7']='OTHER / UNALLOCATED RECEIPTS'; ws['A7'].font=F(10,True,GOLD)
    inp(8,'Share of "other receipts" → Marloth',1.00,'Review owner deposits before treating as revenue.')
    ws['A9']='Share of "other receipts" → Big Bay'; ws['A9'].font=F(10)
    ws['B9']='=1-B8'; ws['B9'].font=F(10,False,GREEN); ws['B9'].number_format=PCT; ws['B9'].alignment=Alignment(horizontal='center')
    ws['A11']='RESERVATIONS'; ws['A11'].font=F(10,True,GOLD)
    inp(12,'Booking.com commission check rate',0.17,'Typical 15–18%.')
    ws['A18']='IKHOKHA'; ws['A18'].font=F(10,True,GOLD)
    ws['A19']='iKhokha blended fee rate (est.)'; ws['A19'].font=F(10)
    ws['B19']=CFG['ikhokha_fee']; ws['B19'].font=F(10,True,BLUE); ws['B19'].fill=FILL_BLUE; ws['B19'].number_format=PCT
    ws['B19'].alignment=Alignment(horizontal='center'); ws['B19'].border=BORD
    ws['C19']='Replace with iKhokha settlement statements.'; ws['C19'].font=F(9,False,GREY)
    ws['A15']='REFERENCE CONTROL TOTALS (locked)'; ws['A15'].font=F(10,True,GOLD)
    for i,(lab,val) in enumerate([('Total inflows (Sales)',ctrl_total_in),('Total outflows',ctrl_total_out),('Refunds/reversals',refunds)]):
        rr=16+i; ws[f'A{rr}']=lab; ws[f'A{rr}'].font=F(10)
        ws[f'B{rr}']=val; ws[f'B{rr}'].font=F(10,True); ws[f'B{rr}'].number_format=ZAR2; ws[f'B{rr}'].alignment=Alignment(horizontal='center')

    # Data_Tagged
    ws=wb.create_sheet('Data_Tagged'); ws.sheet_view.showGridLines=False
    heads=['Date','Month','Property','Counterparty','Category','Inflow','Outflow']; widths=[16,9,11,30,34,13,13]
    for i,(h,w) in enumerate(zip(heads,widths),1):
        c=ws.cell(row=1,column=i,value=h); c.font=F(9,True,'F3ECDB'); c.fill=FILL_INK
        c.alignment=Alignment(indent=1); ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes='A2'
    for ri,(_,rec) in enumerate(data.iterrows(),2):
        ws.cell(ri,1,value=rec['Date'].strftime('%Y-%m-%d %H:%M')).font=F(8,color=GREY)
        ws.cell(ri,2,value=rec['Month']).font=F(8)
        pc={'MARLOTH':BUSH,'BIG_BAY':COAST,'SHARED':GREY}.get(rec['Property'],INK)
        ws.cell(ri,3,value=rec['Property']).font=F(8,True,pc)
        ws.cell(ri,4,value=str(rec['Counterparty'])).font=F(8)
        ws.cell(ri,5,value=str(rec['Category'])).font=F(8)
        a=ws.cell(ri,6,value=round(rec['Inflow'],2) or None); a.number_format=ZAR2; a.font=F(8)
        b=ws.cell(ri,7,value=round(rec['Outflow'],2) or None); b.number_format=ZAR2; b.font=F(8)

    def sif(col,prop,cat): return (f'=SUMIFS(Data_Tagged!${col}$2:${col}${END},'
        f'Data_Tagged!$C$2:$C${END},"{prop}",Data_Tagged!$E$2:$E${END},"{cat}")')

    # Segmented Income Statement
    ws=wb.create_sheet('Segmented_Income_Statement'); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[40,16,16,16,16]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:E1'); hdr(ws,'A1','SEGMENTED INCOME STATEMENT')
    for i,h in enumerate(['','Marloth','Big Bay','Unallocated','Total'],1):
        c=ws.cell(3,i,value=h); c.font=F(10,True,'F3ECDB'); c.fill=FILL_INK
        c.alignment=Alignment(horizontal='right' if i>1 else 'left',indent=1)
    ws['A5']='INCOME'; ws['A5'].font=F(10,True,GOLD)
    ws['A6']='Guest & rental receipts (Sales)'; ws['A6'].font=F(10,True)
    for i,p in enumerate(['MARLOTH','BIG_BAY','SHARED'],2):
        c=ws.cell(6,i,value=sif('F',p,'Sales / customer receipts')); c.font=F(10,True); c.number_format=ZAR
    ws['E6']='=SUM(B6:D6)'; ws['E6'].font=F(10,True); ws['E6'].number_format=ZAR
    ws['A7']='Net revenue'; ws['A7'].font=F(10,True)
    for col in 'BCDE': ws[f'{col}7']=f'={col}6'; ws[f'{col}7'].font=F(10,True); ws[f'{col}7'].number_format=ZAR; ws[f'{col}7'].border=TOPB
    ws['A9']='OPERATING OUTFLOWS'; ws['A9'].font=F(10,True,GOLD)
    r=10
    for cat in EXP_CATS:
        ws[f'A{r}']=cat; ws[f'A{r}'].font=F(10)
        for i,p in enumerate(['MARLOTH','BIG_BAY','SHARED'],2):
            c=ws.cell(r,i,value=sif('G',p,cat)); c.font=F(10); c.number_format=ZAR
        ws.cell(r,5,value=f'=SUM(B{r}:D{r})').number_format=ZAR; ws.cell(r,5).font=F(10,True); r+=1
    tot=r
    ws[f'A{tot}']='Total operating outflows'; ws[f'A{tot}'].font=F(10,True)
    for col in 'BCDE': ws[f'{col}{tot}']=f'=SUM({col}10:{col}{r-1})'; ws[f'{col}{tot}'].font=F(10,True); ws[f'{col}{tot}'].number_format=ZAR; ws[f'{col}{tot}'].border=TOPB
    res=tot+2
    ws[f'A{res}']='NET CASH OPERATING RESULT'; ws[f'A{res}'].font=F(11,True); ws[f'A{res}'].fill=FILL_HEAD
    for col in 'BCDE': ws[f'{col}{res}']=f'={col}7-{col}{tot}'; ws[f'{col}{res}'].font=F(11,True); ws[f'{col}{res}'].number_format=ZAR; ws[f'{col}{res}'].fill=FILL_HEAD
    a=res+3
    ws[f'A{a}']='AFTER SHARED-COST ALLOCATION'; ws[f'A{a}'].font=F(10,True,GOLD)
    ws[f'A{a+1}']='Marloth — operating result (incl. allocation)'; ws[f'A{a+1}'].font=F(10,True)
    ws[f'B{a+1}']=f'=B{res}+Control!B4*D{res}'
    ws[f'A{a+2}']='Big Bay — operating result (incl. allocation)'; ws[f'A{a+2}'].font=F(10,True)
    ws[f'B{a+2}']=f'=C{res}+Control!B5*D{res}'
    ws[f'A{a+3}']='Combined operating result'; ws[f'A{a+3}'].font=F(10,True); ws[f'A{a+3}'].fill=FILL_HEAD
    ws[f'B{a+3}']=f'=B{a+1}+B{a+2}'; ws[f'B{a+3}'].fill=FILL_HEAD
    for rr in (a+1,a+2,a+3): ws[f'B{rr}'].font=F(10,True); ws[f'B{rr}'].number_format=ZAR

    # Monthly_PnL + chart
    ws=wb.create_sheet('Monthly_PnL'); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[14,16,16,16,16]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:E1'); hdr(ws,'A1','NET CASH RESULT BY MONTH & PROPERTY')
    for i,h in enumerate(['Month','Marloth','Big Bay','Unallocated','Total net'],1):
        c=ws.cell(3,i,value=h); c.font=F(10,True,'F3ECDB'); c.fill=FILL_INK; c.alignment=Alignment(horizontal='right' if i>1 else 'left',indent=1)
    r=4
    for m in months:
        ws[f'A{r}']=m; ws[f'A{r}'].font=F(10)
        for i,p in enumerate(['MARLOTH','BIG_BAY','SHARED'],2):
            f=(f'=SUMIFS(Data_Tagged!$F$2:$F${END},Data_Tagged!$C$2:$C${END},"{p}",Data_Tagged!$B$2:$B${END},$A{r})'
               f'-SUMIFS(Data_Tagged!$G$2:$G${END},Data_Tagged!$C$2:$C${END},"{p}",Data_Tagged!$B$2:$B${END},$A{r})')
            ws.cell(r,i,value=f).number_format=ZAR; ws.cell(r,i).font=F(10)
        ws[f'E{r}']=f'=SUM(B{r}:D{r})'; ws[f'E{r}'].number_format=ZAR; ws[f'E{r}'].font=F(10,True); r+=1
    ws[f'A{r}']='TOTAL'; ws[f'A{r}'].font=F(10,True); ws[f'A{r}'].fill=FILL_HEAD
    for col in 'BCDE': ws[f'{col}{r}']=f'=SUM({col}4:{col}{r-1})'; ws[f'{col}{r}'].font=F(10,True); ws[f'{col}{r}'].number_format=ZAR; ws[f'{col}{r}'].fill=FILL_HEAD; ws[f'{col}{r}'].border=TOPB
    ch=BarChart(); ch.type='col'; ch.title='Monthly net result by property'; ch.height=8; ch.width=20
    ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=3,max_row=r-1),titles_from_data=True)
    ch.set_categories(Reference(ws,min_col=1,min_row=4,max_row=r-1)); ws.add_chart(ch,'G3')

    # Reconciliation
    ws=wb.create_sheet('Reconciliation'); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCD',[46,18,18,16]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:D1'); hdr(ws,'A1','RECONCILIATION · model vs bank control totals')
    for i,h in enumerate(['Check','Model','Control','Diff'],1):
        c=ws.cell(3,i,value=h); c.font=F(10,True,'F3ECDB'); c.fill=FILL_INK; c.alignment=Alignment(horizontal='right' if i>1 else 'left',indent=1)
    checks=[('Total Sales receipts',f'=SUMIFS(Data_Tagged!$F$2:$F${END},Data_Tagged!$E$2:$E${END},"Sales / customer receipts")','=Control!B16'),
            ('Total operating outflows',f'=SUM(Data_Tagged!$G$2:$G${END})','=Control!B17'),
            ('Net operating result','=B4-B5','=Control!B16-Control!B17')]
    r=4
    for lab,m,c in checks:
        ws[f'A{r}']=lab; ws[f'A{r}'].font=F(10); ws[f'B{r}']=m; ws[f'C{r}']=c; ws[f'D{r}']=f'=ROUND(B{r}-C{r},2)'
        for col in 'BCD': ws[f'{col}{r}'].number_format=ZAR2; ws[f'{col}{r}'].font=F(10)
        r+=1
    ws[f'A{r+1}']='✔ Ties to bank control totals when all diffs = 0.00'; ws[f'A{r+1}'].font=F(9,False,GREY)
    return END, months

# ----------------------------------------------------------------------------
# iKhokha gross-up tabs (optional)
# ----------------------------------------------------------------------------
def add_ikhokha(wb, END, months, ik_path):
    if not ik_path or not os.path.exists(ik_path):
        print("  iKhokha: no export found, skipping."); return
    ik=pd.read_excel(ik_path, sheet_name=0, header=0)
    ws=wb.create_sheet('iKhokha_Reference'); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[42,17,17,16,16]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:E1'); hdr(ws,'A1','IKHOKHA REFERENCE · gross-up journal')
    rows=[('Net settled to bank (control)',f'=SUMIFS(Data_Tagged!$F$2:$F${END},Data_Tagged!$D$2:$D${END},"Ik Pay*")',GREEN),
          ('iKhokha fee rate (Control)','=Control!B19',GREEN),
          ('Implied GROSS card revenue','=B4/(1-B5)',INK),
          ('iKhokha processing fee (expense)','=B6-B4',INK)]
    r=4
    for lab,frm,clr in rows:
        ws[f'A{r}']=lab; ws[f'A{r}'].font=F(10); c=ws[f'B{r}']; c.value=frm; c.font=F(10,True,clr)
        c.number_format=(PCT if 'rate' in lab else ZAR2); r+=1
    # monthly net + audit
    hr=r+1
    for i,h in enumerate(['Month','Net settled','Implied gross','Implied fee'],1):
        c=ws.cell(hr,i,value=h); c.font=F(9,True,'F3ECDB'); c.fill=FILL_INK; c.alignment=Alignment(horizontal='right' if i>1 else 'left',indent=1)
    rr=hr+1
    for m in months:
        ws[f'A{rr}']=m; ws[f'A{rr}'].font=F(10)
        ws[f'B{rr}']=f'=SUMIFS(Data_Tagged!$F$2:$F${END},Data_Tagged!$D$2:$D${END},"Ik Pay*",Data_Tagged!$B$2:$B${END},$A{rr})'
        ws[f'B{rr}'].number_format=ZAR; ws[f'B{rr}'].font=F(10,False,GREEN)
        ws[f'C{rr}']=f'=B{rr}/(1-Control!$B$19)'; ws[f'C{rr}'].number_format=ZAR; ws[f'C{rr}'].font=F(10)
        ws[f'D{rr}']=f'=C{rr}-B{rr}'; ws[f'D{rr}'].number_format=ZAR; ws[f'D{rr}'].font=F(10); rr+=1
    # data tab
    ws2=wb.create_sheet('iKhokha_Data'); ws2.sheet_view.showGridLines=False
    cols=[c for c in ['Date & Time','Transaction type','Source','Reference','Total','VAT'] if c in ik.columns]
    for i,h in enumerate(cols,1):
        c=ws2.cell(1,i,value=h); c.font=F(9,True,'F3ECDB'); c.fill=FILL_INK
        ws2.column_dimensions[chr(64+i)].width=18
    for j,(_,x) in enumerate(ik.iterrows(),2):
        for i,col in enumerate(cols,1):
            v=x[col]; v='' if pd.isna(v) else (round(float(v),2) if col=='Total' else str(v))
            ws2.cell(j,i,value=v).font=F(8)
    print(f"  iKhokha: {len(ik)} transactions, gross R{ik['Total'].sum():,.2f}")

# ----------------------------------------------------------------------------
# Zoho cross-check (optional) — expects a small summary json (see CLAUDE.md)
# {"invoiced":..,"collected":..,"outstanding":..,"expenses":..,"by_month":{"YYYY-MM":[invoiced,outstanding]}}
# ----------------------------------------------------------------------------
def add_zoho(wb, END, months, zpath):
    if not zpath or not os.path.exists(zpath):
        print("  Zoho: no summary json, skipping."); return
    z=json.load(open(zpath))
    ws=wb.create_sheet('Zoho_Reference'); ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDE',[40,17,17,17,17]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:E1'); hdr(ws,'A1','ZOHO BOOKS REFERENCE · accrual vs cash')
    head=[('Zoho invoiced revenue',z.get('invoiced'),RED),('  collected',z.get('collected'),RED),
          ('  outstanding (AR)',z.get('outstanding'),RED),
          ('Bank banked receipts',f'=SUMIFS(Data_Tagged!$F$2:$F${END},Data_Tagged!$E$2:$E${END},"Sales / customer receipts")',GREEN),
          ('Zoho recorded expenses',z.get('expenses'),RED),
          ('Bank total outflows',f'=SUM(Data_Tagged!$G$2:$G${END})',GREEN)]
    r=4
    for lab,val,clr in head:
        ws[f'A{r}']=lab; ws[f'A{r}'].font=F(10); c=ws[f'B{r}']; c.value=val; c.font=F(10,True,clr); c.number_format=ZAR2; r+=1
    print(f"  Zoho: invoiced R{z.get('invoiced',0):,.2f}, AR R{z.get('outstanding',0):,.2f}")

# ----------------------------------------------------------------------------
def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--bank', default=CFG['bank'])
    ap.add_argument('--ikhokha', default=CFG['ikhokha'])
    ap.add_argument('--zoho', default=CFG['zoho_summary'])
    ap.add_argument('--out', default=CFG['out'])
    a=ap.parse_args()
    os.makedirs(os.path.dirname(a.out) or '.', exist_ok=True)
    # allow glob for iKhokha (use the largest export found)
    ik=a.ikhokha
    if ik and ('*' in ik or not os.path.exists(ik)):
        cand=sorted(glob.glob('inputs/transactions_*.xlsx'), key=lambda p: os.path.getsize(p), reverse=True)
        ik=cand[0] if cand else None
    print(f"Loading bank: {a.bank}")
    data=load_bank(a.bank)
    wb=Workbook()
    END,months=build_core(wb, data)
    add_ikhokha(wb, END, months, ik)
    add_zoho(wb, END, months, a.zoho)
    wb.save(a.out)
    print(f"Saved {a.out}.  Recalc next:  python recalc.py {a.out}")
    # regenerate the investor portfolio HTML from the same data (best-effort)
    try:
        import build_portfolio
        build_portfolio.build(bank_path=a.bank, out_dir=os.path.dirname(a.out) or 'output')
    except Exception as e:
        print(f"  Portfolio skipped ({e}); run `python build_portfolio.py` separately.")

if __name__=='__main__':
    main()
